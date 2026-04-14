import openpyxl

def search_in_workbook(filepath, keywords_str, match_exact, search_cols_str, search_rows_str, search_sheets_str):
    """
    Searches for keywords in the given Excel workbook and returns a list of results.
    Filters by cols, rows, and sheets if provided.
    
    Returns a list of dicts:
    [
        {
            "filepath": absolute path to file,
            "sheet_name": sheet name where found,
            "cell": coordinate (e.g., 'A1'),
            "row": row index,
            "column": column index,
            "text": The full text content of the cell,
            "keyword": The exact keyword that was matched
        }, ...
    ]
    """
    results = []
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        keywords = [k.strip() for k in keywords_str.split('|') if k.strip()]
        keywords_lower = [k.lower() for k in keywords]
        
        target_sheets = [s.strip().lower() for s in search_sheets_str.split('|') if s.strip()]
        
        # Convert target_cols to column indices (1-indexed based on openpyxl conventions, A=1, B=2)
        target_cols = [c.strip().upper() for c in search_cols_str.split('|') if c.strip()]
        from openpyxl.utils.cell import column_index_from_string
        target_col_indices = []
        for c in target_cols:
            try:
                target_col_indices.append(column_index_from_string(c))
            except:
                pass
                
        target_rows = []
        for r in search_rows_str.split('|'):
            if r.strip().isdigit():
                target_rows.append(int(r.strip()))
                
        for sheet_name in wb.sheetnames:
            if target_sheets and sheet_name.lower() not in target_sheets:
                continue
                
            ws = wb[sheet_name]
            sheet_results = []
            
            for row in ws.iter_rows():
                # Check row filter first
                if row and len(row) > 0:
                    current_row_idx = row[0].row
                    if target_rows and current_row_idx not in target_rows:
                        continue
                        
                for cell in row:
                    if target_col_indices and cell.column not in target_col_indices:
                        continue
                        
                    if cell.value is not None:
                        # Convert to string to handle numeric/date cell types
                        cell_val_str = str(cell.value)
                        
                        matched_kws = []
                        if match_exact:
                            cell_lower = cell_val_str.lower()
                            for i, kw_lower in enumerate(keywords_lower):
                                if cell_lower == kw_lower:
                                    if keywords[i] not in matched_kws:
                                        matched_kws.append(keywords[i])
                        else:
                            cell_lower = cell_val_str.lower()
                            for i, kw_lower in enumerate(keywords_lower):
                                if kw_lower in cell_lower:
                                    if keywords[i] not in matched_kws:
                                        matched_kws.append(keywords[i])
                                
                        if matched_kws:
                            sheet_results.append({
                                "filepath": filepath,
                                "sheet_name": sheet_name,
                                "cell": cell.coordinate,
                                "row": cell.row,
                                "column": cell.column,
                                "text": cell_val_str,
                                "keywords": matched_kws,
                                "keyword": "|".join(matched_kws)
                            })
                            
            # Sort by row then column
            sheet_results.sort(key=lambda x: (x['row'], x['column']))
            results.extend(sheet_results)
            
        wb.close()
        
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        
    return results
